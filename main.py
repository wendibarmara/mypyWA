# importing openpyxl module
import openpyxl

# Give the location of the file
path = "C:\\mypyWA\\data\\sample.xlsx"
# workbook object is created
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
m_row = sheet_obj.max_row
# Loop will print all values
# of first column
# for i in range(2, m_row + 1):
#    cell_obj = sheet_obj.cell(row=i, column=3)
#    if cell_obj.value == "Hasan":
#        print("ini hasan")
#   else:
#      print(cell_obj.value)
